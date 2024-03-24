import camelot

from openpyxl import Workbook
from datetime import datetime

import json

import fitz #If the library is not installed, you can install it by typing the following command in the command prompt: pip install PyMuPDF

'''If you do not have Tesseract OCR installed, you can refer to the official website of Tesseract OCR 
for instructions on how to install it: https://github.com/tesseract-ocr/tesseract'''
import pytesseract #If the library is not installed, you can install it by typing the following command in the command prompt: pip install pytesseract
from PIL import Image

import os
import shutil


#You need to enter the file path of the folder where the tax plate pdfs are located -- Example: C:\User\Desktop\input pdf
pdf_folder_path = r""

os.makedirs(r"read_pdf_project", exist_ok=True)
os.makedirs(r"read_pdf_project\read_pdf_result", exist_ok=True)
os.makedirs(r"read_pdf_project\pdf_images", exist_ok=True)

wb = Workbook()
ws = wb.active

headers = ["ADI SOYADI", "TICARET UNVANI", "IS YERI ADRESI",
           "VERGI TURU", "VERGI DAIRESI", "VERGI KIMLIK NO",
           "TC KIMLIK NO", "ISE BASLAMA TARIHI", "FAALIYET KOD VE ADLARI"]

for col_index, header_value in enumerate(headers, start=1):
    ws.cell(row=1, column=col_index, value=header_value)


pdf_file_names = os.listdir(pdf_folder_path)

for pdf_count, pdf_file in enumerate(pdf_file_names):
    pdf_path = pdf_folder_path + "\\" + pdf_file

    tables = camelot.read_pdf(pdf_path, flavor='stream', columns=['218,490,570'], row_tol=30)

    tables[0].to_json(r'read_pdf_project\read_pdf_result\result_'+pdf_file+'.json')


    with open(r'read_pdf_project\read_pdf_result\result_'+pdf_file+'.json', encoding='utf-8') as f:
        data = json.load(f)


    
    pdf_dosyasi = fitz.open(pdf_path)
    images = []

    pdf_page = pdf_dosyasi.load_page(0)

    for pdf_image_no, pdf_image in enumerate(pdf_page.get_images(full=True)):
        image_pixel = pdf_dosyasi.extract_image(pdf_image[0])
        images.append(image_pixel)

    for pdf_image_no, pdf_image in enumerate(images, start=1):
        pdf_image_name = f"read_pdf_project\pdf_images\pdf_image_{pdf_image_no}.png"
        with open(pdf_image_name, "wb") as f:
            f.write(pdf_image["image"])
    
    image_file_names = os.listdir("read_pdf_project\pdf_images")
    for image_file in image_file_names:
        image_path = "read_pdf_project\pdf_images\\" + image_file

        image = Image.open(image_path)

        text = pytesseract.image_to_string(image)

        vergi_kimlik_no = ''
        if text != '' and text.split()[0].isdigit():
            vergi_kimlik_no = text.split()[0]
            break
    
    total_value = [data[1]["1"].replace('\n',' '), data[2]["1"].replace('\n',' '),
                   data[3]["1"].replace('\n',' '), data[4]["1"].replace('\n',' '),
                   data[1]["3"].replace('\n',' '), vergi_kimlik_no,
                   data[3]["3"].replace('\n',' '), data[4]["3"].replace('\n',' '),
                   data[5]["2"].replace('\n',' ')+data[5]["1"].replace('\n',' ')]


    for col_index, cell_value in enumerate(total_value, start=1):
        ws.cell(row=pdf_count+2, column=col_index, value=cell_value)

    [os.remove(os.path.join("read_pdf_project\pdf_images", file)) for file in os.listdir("read_pdf_project\pdf_images")]

wb.save(r"read_pdf_project\result_"+datetime.today().strftime('%Y-%m-%d')+".xlsx")

shutil.rmtree(r"read_pdf_project\read_pdf_result")
shutil.rmtree(r"read_pdf_project\pdf_images")
